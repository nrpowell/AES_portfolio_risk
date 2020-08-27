from argparse import ArgumentParser
from datetime import datetime
from itertools import repeat

from constants import minimum_contract_date, maximum_contract_date

import csv
import math
import openpyxl
import os
import random
import sys
import time
import xlrd

import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt 

def print_step_figs(matx, xvals, path):
    plt.clf()
    for i in range(0, matx.shape[0]):
        plt.step(xvals, list(matx[i]), where='mid')
    plt.savefig(path)

def make_aes_graphs(simulated_prices_dict, num_simulations, directory):
    plt.rc('xtick', labelsize=5) 
    plt.rc('ytick', labelsize=5)
    plt.rc('font', size=5)
    plt.rc('axes', labelsize=5)
    plt.rc('axes', titlesize=5)
    plt.rcParams["figure.figsize"] = (6,4)

    if not os.path.isdir(f"{directory}/step_sims"):
        os.makedirs(f"{directory}/step_sims")

    for key, value in simulated_prices_dict.items():
        contracts = value.keys()

        ## A kind of complicated way to sort the list of string dates
        sorted_keys_tuples = sorted(list(map(datetime.strptime, contracts, repeat("%B%Y", times=len(contracts)))))
        sorted_keys = list(map(datetime.strftime, sorted_keys_tuples, repeat("%B%Y", times=len(sorted_keys_tuples))))

        price_curve_1day = np.zeros((num_simulations, len(sorted_keys)))
        price_curve_5days = np.zeros((num_simulations, len(sorted_keys)))
        price_curve_20days = np.zeros((num_simulations, len(sorted_keys)))
        for i in range(0, len(sorted_keys)):
            contract = sorted_keys[i]
            all_simulations_prices = value[contract]

            for j in range(0, len(all_simulations_prices)):
                one_simulation_prices = all_simulations_prices[j]

                price_curve_1day[j, i]      = one_simulation_prices[1]
                price_curve_5days[j, i]     = one_simulation_prices[5]
                price_curve_20days[j, i]    = one_simulation_prices[20]


        print_step_figs(price_curve_1day, sorted_keys, f"{directory}/step_sims/{key}_1day.png")
        print_step_figs(price_curve_5days, sorted_keys, f"{directory}/step_sims/{key}_5days.png")
        print_step_figs(price_curve_20days, sorted_keys, f"{directory}/step_sims/{key}_20days.png")

def graph_price_histories(all_simulations_prices, currency_date, directory):
    plt.clf()
    counter = 0
    for sim in all_simulations_prices:
        plt.plot(sim)
        counter += 1
    print(f"About to write file {currency_date}")
    if not os.path.isdir(f"{directory}/sims"):
        os.makedirs(f"{directory}/sims")
    plt.savefig(f"{directory}/sims/{currency_date}")

def plot_histograms(all_positions_1day, all_positions_5day, all_positions_20day, directory):
    if not os.path.isdir(f"{directory}/histograms"):
        os.makedirs(f"{directory}/histograms")

    for date, pos1 in all_positions_1day.items():
        if not os.path.isdir(f"{directory}/histograms/1day"):
            os.makedirs(f"{directory}/histograms/1day")
        plt.clf()
        fifth_percentile = pos1[int(len(pos1)/20)]
        plt.hist(pos1, 30, edgecolor='grey', color = "skyblue")
        plt.plot(fifth_percentile, 0, 'ro')
        plt.annotate(f"5% = ${int(fifth_percentile)}", (fifth_percentile, 0), textcoords="offset points", xytext=(0,10), ha='center', size = 8, weight='bold')
        path = f"{directory}/histograms/1day/{date}_1.png"
        plt.xlabel("Dollar value of portfolio")
        plt.ylabel("Number of simulations (out of 1000)")
        plt.title(f"{date}, 1 holding day")
        plt.savefig(path)

    for date, pos5 in all_positions_5day.items():
        if not os.path.isdir(f"{directory}/histograms/5day"):
            os.makedirs(f"{directory}/histograms/5day")
        plt.clf()
        fifth_percentile = pos5[int(len(pos5)/20)]
        plt.hist(pos5, 30, edgecolor='grey', color = "skyblue")
        plt.plot(fifth_percentile, 0, 'ro')
        plt.annotate(f"5% = ${int(fifth_percentile)}", (fifth_percentile, 0), textcoords="offset points", xytext=(0,10), ha='center', size = 8, weight='bold')
        path = f"{directory}/histograms/5day/{date}_5.png"
        plt.xlabel("Dollar value of portfolio")
        plt.ylabel("Number of simulations (out of 1000)")
        plt.title(f"{date}, 5 holding days")
        plt.savefig(path)


    for date, pos20 in all_positions_20day.items():
        if not os.path.isdir(f"{directory}/histograms/20day"):
            os.makedirs(f"{directory}/histograms/20day")
        plt.clf()
        fifth_percentile = pos20[int(len(pos20)/20)]
        plt.hist(pos20, 30, edgecolor='grey', color = "skyblue")
        plt.plot(fifth_percentile, 0, 'ro')
        plt.annotate(f"5% = ${int(fifth_percentile)}", (fifth_percentile, 0), textcoords="offset points", xytext=(0,10), ha='center', size = 8, weight='bold')
        path = f"{directory}/histograms/20day/{date}_20.png"
        plt.xlabel("Dollar value of portfolio")
        plt.ylabel("Number of simulations (out of 1000)")
        plt.title(f"{date}, 20 holding days")
        plt.savefig(path)


def collect_exposure_data(directory):
    wb = openpyxl.load_workbook(f'{directory}/{directory}_exposures.xlsx')
    sheet = wb['Sheet1']

    ## First, we need to determine the start and end dates (columns) of the data, based on the contracts that we have in use...
    ## First row is contract dates
    ## Cell (1,1) is irrelevant for our purposes, so we start at 2
    min_column = 0
    max_column = sheet.max_column
    column_to_datestring_dict = {}
    for c in range(2, sheet.max_column):
        vl = sheet.cell(1, c).value
        column_to_datestring_dict[c] = vl.strftime("%B%Y")
        if min_column == 0:
            if datetime(*minimum_contract_date) == vl:
                min_column = c
        elif max_column == sheet.max_column:
            if datetime(*maximum_contract_date) == vl:
                max_column = c

    ## Now, we go through each row in the exposures dataset, get the currency-date contract string, and store its exposure in a dict
    curr_date_exposure_dict = {}
    for r in range(2, sheet.max_row + 1):
        raw_curr_string = sheet.cell(r, 1).value
        if directory == "FX":
            curr_string = "FX_" + raw_curr_string
        elif directory == "IR":
            curr_string = "IR_" + raw_curr_string
        else:
            pass

        for c in range(min_column, max_column):
            contract_date_string = column_to_datestring_dict[c]
            exposure = sheet.cell(r, c).value
            full_string = curr_string + "_" + contract_date_string

            if contract_date_string in curr_date_exposure_dict:
                running_dict = curr_date_exposure_dict[contract_date_string]
                running_dict[curr_string] = exposure
                curr_date_exposure_dict[contract_date_string] = running_dict
            else:
                new_dict = {curr_string: exposure}
                curr_date_exposure_dict[contract_date_string] = new_dict

            #curr_date_exposure_dict[full_string] = exposure

    return curr_date_exposure_dict


def calculate_var_for_holding_day(simulated_prices_dict, curr_date_exposure_dict, num_simulations, contract_to_rep_dict, holding_days):
    all_portfolio_positions = {}
    for date, currency_dict in curr_date_exposure_dict.items():
        exposures_dict = {}
        for currency, exposure in currency_dict.items():
            if exposure:
                rep = contract_to_rep_dict[currency + "_" + date]
                if rep in exposures_dict:
                    running_exposure = exposures_dict[rep]
                    exposures_dict[rep] = running_exposure + exposure
                else:
                    exposures_dict[rep] = exposure

        #print(f"List of top level keys: {simulated_prices_dict.keys()}")
        #print(f"Exposures dict items: {exposures_dict.keys()}; sim prices dict items: {simulated_prices_dict[date].keys()}")

        ## We filter down the prices dict to only include those contracts in the exposures dict
        filtered_prices_dict = {}
        for rep, exposure in exposures_dict.items():
            rep_date = rep.split('_')[2]
            rep_currency = rep.split(rep_date)[0][:-1]
            price_list = simulated_prices_dict[rep_date][rep_currency]
            filtered_prices_dict[rep_currency + "_" + rep_date] = price_list

        #filtered_prices_dict = dict(filter(lambda elem: elem[0] + "_" + date in exposures_dict, simulated_prices_dict[date].items()))

        portfolio_positions = []
        for i in range(0, num_simulations-1):
            position = 0.0
            for contract_date, price_list in filtered_prices_dict.items():
                exposure = exposures_dict[contract_date]
                #print(f"Exposure is {exposure}")
                ## TODO: here is where we can store the original price
                simulation_price_to_use = price_list[i][holding_days]
                original_price = price_list[i][0]
                if not "ARS" in contract_date:
                    if contract_date.startswith("FX"):
                        position += (exposure / simulation_price_to_use)
                    else:
                        ## TODO: Here is where we will distinguish between floating and fixed rates
                        position += (exposure * (1.0 + simulation_price_to_use))

            portfolio_positions.append(position)
           #print(f"Position is {position}\n\n")

        portfolio_positions.sort()
        all_portfolio_positions[date] = portfolio_positions

    return all_portfolio_positions


def determine_var(simulated_prices_dict, curr_date_exposure_dict, num_simulations, directory):
    ## First, we have to create a dict that links every currency contract to the currency contract whose price list we will be using
    keys_set = set()
    for key, value in simulated_prices_dict.items():
        for subkey in value.keys():
            keys_set.add(subkey + "_" + key)
    contract_to_rep_dict = {}
    fp = os.getcwd() + f"/{directory}/Clusters.csv"
    reader = csv.reader(open(fp,'r'))
    for row in reader:
        running_list = []
        representative = ""
        for i in range(0, len(row)):
            contract = row[i]
            if contract in keys_set:
                representative = contract
            else:
                running_list.append(contract)
        assert representative != ""
        for c in running_list:
            contract_to_rep_dict[c] = representative
        contract_to_rep_dict[representative] = representative


    ## Next we determine all potential portfolio positions at 1, 5, and 20 holding days
    all_positions_1day = calculate_var_for_holding_day(simulated_prices_dict, curr_date_exposure_dict, num_simulations, contract_to_rep_dict, 1)
    all_positions_5day = calculate_var_for_holding_day(simulated_prices_dict, curr_date_exposure_dict, num_simulations, contract_to_rep_dict, 5)
    all_positions_20day = calculate_var_for_holding_day(simulated_prices_dict, curr_date_exposure_dict, num_simulations, contract_to_rep_dict, 20)

    plot_histograms(all_positions_1day, all_positions_5day, all_positions_20day, directory)



"""def determine_var(simulated_prices_dict, num_simulations, directory):
    wb = openpyxl.load_workbook(f'{directory}/{directory}_exposures.xlsx')
    sheet = wb['Sheet1']

    ## First, we need to determine the start and end dates (columns) of the data, based on the contracts that we have in use...
    ## First row is contract dates
    ## Cell (1,1) is irrelevant for our purposes, so we start at 2
    #minimum_contract_date = datetime(*minimum_contract_date)
    #maximum_contract_date = datetime(*maximum_contract_date)
    min_column = 0
    max_column = sheet.max_column
    column_to_datestring_dict = {}
    #print(f"minimum_contract_date is {minimum_contract_date}; maximum_contract_date is {maximum_contract_date}")
    for c in range(2, sheet.max_column):
        vl = sheet.cell(1, c).value
        #contract_date = datetime.strptime(vl.strftime("%d-%m"), "%y-%m")
        column_to_datestring_dict[c] = vl.strftime("%B%Y")
        if min_column == 0:
            #print(f"Constant minimum date: {datetime(*minimum_contract_date)}, contract date: {contract_date} and raw cell value is {vl}")
            if datetime(*minimum_contract_date) == vl:#contract_date:
                min_column = c
        elif max_column == sheet.max_column:
            if datetime(*maximum_contract_date) == vl:#contract_date:
                max_column = c

    ## Next, we have to create a dict that links every currency contract to the currency contract whose price list we will be using
    keys_set = set()
    for key, value in simulated_prices_dict.items():
        for subkey in value.keys():
            keys_set.add(key + "_" + subkey)
    contract_to_rep_dict = {}
    fp = os.getcwd() + f"/{directory}/Clusters.csv"
    reader = csv.reader(open(fp,'r'))
    for row in reader:
        running_list = []
        representative = ""
        for i in range(0, len(row)):
            contract = row[i]
            if contract in keys_set:
                representative = contract
            else:
                running_list.append(contract)
        print(f"\n\n\n")
        assert representative != ""
        for c in running_list:
            contract_to_rep_dict[c] = representative
        contract_to_rep_dict[representative] = representative

    ## We will be writing the results to a new Excel workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    column_width = 25

    current_row = 1

    ## Now, we go through each row in the exposures dataset and determine the 5th percentile loss for each contract after 1, 3, and 20 days, storing the results in a dict
    for r in range(2, sheet.max_row):
        curr_string = "FX_USD" + sheet.cell(r, 1).value

        ## Initialize the 1-column text
        column_letter = new_sheet.cell(current_row, 1).column_letter
        new_sheet.cell(current_row, 1).value = curr_string
        new_sheet.cell(current_row + 1, 1).value = "Exposure"
        new_sheet.cell(current_row + 2, 1).value = "USD value (day 0)"
        new_sheet.cell(current_row + 3, 1).value = "1-day 5\% risk"
        new_sheet.cell(current_row + 4, 1).value = "5-day 5\% risk"
        new_sheet.cell(current_row + 5, 1).value = "20-day 5\% risk"
        new_sheet.column_dimensions[column_letter].width = column_width

        reps_price_lists = simulated_prices_dict[curr_string]

        current_column = 2
        for c in range(min_column, max_column):
            contract_date_string = column_to_datestring_dict[c]
            full_string = curr_string + "_" + contract_date_string
            representative = contract_to_rep_dict[full_string]

            ## Chop off the currency name
            price_list = reps_price_lists[representative[10:]]
            
            exposure = sheet.cell(r, c).value
            original_price = price_list[0][0]

            prices_1day = []
            prices_5day = []
            prices_20day = []

            for lst in price_list:
                prices_1day.append(lst[1])
                prices_5day.append(lst[5])
                prices_20day.append(lst[20])

            prices_1day.sort()
            prices_5day.sort()
            prices_20day.sort()

            ## If the exposure is positive (profit coming in), AES benefits from a low exchange rate
            ## If the exposure is negative (losses accruing), AES benefits from a high exchange rate
            ## Therefore, the 5th percentile worst-case scenario will be the 950th element in the sorted list with positive exposure, and the 50th element in the sorted list with negative exposure

            risk_price_1day = prices_1day[50] if exposure < 0 else prices_1day[950]
            risk_price_5day = prices_5day[50] if exposure < 0 else prices_5day[950]
            risk_price_20day = prices_20day[50] if exposure < 0 else prices_20day[950]

            column_letter = new_sheet.cell(current_row, current_column).column_letter
            new_sheet.cell(current_row, current_column).value = contract_date_string
            new_sheet.cell(current_row + 1, current_column).value = exposure if exposure > 0 else f"({exposure})"
            new_sheet.cell(current_row + 2, current_column).value = exposure / original_price if exposure > 0 else f"({exposure / original_price})"
            new_sheet.cell(current_row + 3, current_column).value = exposure / risk_price_1day if exposure > 0 else f"({exposure / risk_price_1day})"
            new_sheet.cell(current_row + 4, current_column).value = exposure / risk_price_5day if exposure > 0 else f"({exposure / risk_price_5day})"
            new_sheet.cell(current_row + 5, current_column).value = exposure / risk_price_20day if exposure > 0 else f"({exposure / risk_price_20day})"
            new_sheet.column_dimensions[column_letter].width = column_width
            current_column += 1

        current_row += 8


    ## Finally, we go through the columns one last time, one at a time, and calculate 5% exposure rates for the portfolio as a whole at each contract date
    ## A lot of this code is reused from above. This isn't the *best* software practice, but it does the job
    for c in range(min_column, max_column):
        contract_date_string = column_to_datestring_dict[c]

        exposures_dict = {}
        original_price_dict = {}
        price_list_dict = {}
        for r in range(2, sheet.max_row):
            curr_string = "FX_USD" + sheet.cell(r, 1).value

            reps_price_lists = simulated_prices_dict[curr_string]

            full_string = curr_string + "_" + contract_date_string
            representative = contract_to_rep_dict[full_string]

            ## Chop off the currency name
            price_list = reps_price_lists[representative[10:]]

            original_price = price_list[0][0]
            
            exposure = sheet.cell(r, c).value

            exposures_dict[full_string] = exposure
            original_price_dict[full_string] = original_price
            price_list_dict[full_string] = price_list

        ## Get day-0 USD value of portfolio
        usd0_value = 0.0
        for key, value in exposures_dict.items():
            original_price = original_price_dict[key]
            usd0_value += (value / original_price)

        key_strings = price_list_dict.keys()
        potential_exposures = []
        for s in range(0, num_simulations-1):
            total_exposure_one_sim = 0.0
            for key_string in key_strings:
                price_list = price_list_dict[key_string]
                exposure = exposures_dict[key_string]

                #print(f"S is {s}, but the length of the price list is {len(price_list)}")
                price_1_day = price_list[s][1]
                total_exposure_one_sim += (exposure / price_1_day)
            potential_exposures.append(total_exposure_one_sim)
        potential_exposures.sort()
        fifth_percentile = potential_exposures[50]
        print(f"Fifth percentile for portfolio at {contract_date_string} is: {fifth_percentile}; median is {potential_exposures[500]}")


    new_wb.save(f"{directory}/simulated_VaRs.xlsx")"""


def get_raw_prices(forwards_directory, output_type, is_combined):
    simulation_directory = "combined" if is_combined else forwards_directory
    simulations_dir = f"{simulation_directory}/pcas"
    fps = [f"pcas/{fl}" for fl in os.listdir(simulations_dir) if fl.endswith(".csv")]

    ## To get a set of the contract names (of the cluster representatives), we take the list of sub-directory files from above and just chop off "_cluster.csv" from the end - which is 12 characters from the end - and the directory name from the beginning.
    set_contract_names = set([fp[5:-12] for fp in fps])

    ## Now, we go through the original csv, line by line. If the contract currency and date line up with a set member, we take the price, convert the simulation file log returns to pure prices, and do the plot. Then we remove the element from the set. We are only looking for the *first* price for each contract+date in the csv; that is, the most recent price. This acts as our baseline from which we derive all other prices.
    fp = os.getcwd() + f"/{forwards_directory}/historical_forwards_new.xlsx"
    wb = xlrd.open_workbook(fp)

    sheet = wb.sheet_by_index(0)
    col_currency    = 0
    col_setdate     = 4
    col_contract    = 5
    col_price       = 7

    ## To be changed later
    num_simulations = 0

    simulated_prices_dict = {}
    for i in range(1, sheet.nrows):
        if not set_contract_names:
            break

        tup_contract_date = xlrd.xldate_as_tuple(sheet.cell_value(i, col_contract), wb.datemode)
        dt = datetime(*tup_contract_date)
        month_year = dt.strftime("%B%Y")

        currency_to_match = sheet.cell_value(i, col_currency)
        if forwards_directory == "FX":
            currency_to_match = currency_to_match.replace('USD', '')
        currency_date = currency_to_match + "_" + month_year
        if currency_date in set_contract_names:
            dta = os.getcwd() + f"/{simulation_directory}/pcas/{currency_date}_cluster.csv"
            df = pd.read_csv(dta)

            original_price = sheet.cell_value(i, col_price)
            if currency_to_match == "FX_USDEUR":
                original_price = (1.0 / original_price)

            num_simulations = df.shape[1]
            all_simulations_prices = []
            for j in range(1, df.shape[1]):
                one_simulation_prices = []
                one_simulation_prices.append(original_price)
                for k in range(0, df.shape[0]):
                    raw_return = df.iloc[k, j]
                    curr_price = raw_return * original_price
                    one_simulation_prices.append(curr_price)

                all_simulations_prices.append(one_simulation_prices)

            ## 1 --> this will output a graph of each cluster's set of simulations as they evolve over time
            if output_type == 1:
                graph_price_histories(all_simulations_prices, currency_date, forwards_directory)

            ## Stores all simulated prices in a dict with each key being one currency, and the value being a dict storing the simulated prices for each cluster representative of that currency

            #print(f"Contract: {currency_to_match}_{month_year}")
            if month_year not in simulated_prices_dict:
                new_sub_dict = {currency_to_match: all_simulations_prices}
                simulated_prices_dict[month_year] = new_sub_dict
            else:
                sub_dict = simulated_prices_dict[month_year]
                sub_dict[currency_to_match] = all_simulations_prices
                simulated_prices_dict[month_year] = sub_dict

            set_contract_names.remove(currency_date)



    ## 2 --> this will output a graph similar to that used by AES to represent price curves at 1, 5, and 20 holding days
    if output_type == 2:
        make_aes_graphs(simulated_prices_dict, num_simulations, forwards_directory)

    ## 3 --> return the simulated prices dict to then be used however we want
    return (simulated_prices_dict, num_simulations)


## This merges nested dicts (dicts whose values are dicts) whose top-level keys may be the same
## This also assumes the second-level dict keys are NEVER the same across the two dicts
def custom_dict_merge(dict1, dict2):
    dict3 = {}
    for key1, val1 in dict1.items():
        if key1 in dict2:
            val2 = dict2[key1]

            ## Here we assume that the keys will never be the same across the two dicts
            val3 = {**val1, **val2}
            dict3[key1] = val3
            del dict2[key1]

        else:
            dict3[key1] = val1

    ## This should be all keys that are not shared with dict1
    for key2, val2 in dict2.items():
        dict3[key2] = val2

    return dict3


def main(directory, output_type):
    if output_type == 3:
        if directory != "combined":
            (simulated_prices_dict, num_simulations) = get_raw_prices(directory, output_type, False)
            curr_date_exposure_dict = collect_exposure_data(directory)
            determine_var(simulated_prices_dict, curr_date_exposure_dict, num_simulations, directory)
        else:
            (simulated_prices_dict_FX, num_simulations_FX) = get_raw_prices("FX", output_type, True)
            (simulated_prices_dict_IR, num_simulations_IR) = get_raw_prices("IR", output_type, True)
            assert(num_simulations_FX == num_simulations_IR)
            simulated_prices_dict = custom_dict_merge(simulated_prices_dict_FX, simulated_prices_dict_IR) #{**simulated_prices_dict_FX, **simulated_prices_dict_IR}

            curr_date_exposure_dict_FX = collect_exposure_data("FX")
            curr_date_exposure_dict_IR = collect_exposure_data("IR")
            curr_date_exposure_dict = custom_dict_merge(curr_date_exposure_dict_FX, curr_date_exposure_dict_IR)

            #print(f"The keys of the combined SIMULATION dict are {set(simulated_prices_dict.keys())}")
            #print(f"The keys of the combined EXPOSURE dict are {set(curr_date_exposure_dict.keys())}")
            #print(f"The difference between the two is {set(simulated_prices_dict.keys()) - set(curr_date_exposure_dict.keys())}")

            determine_var(simulated_prices_dict, curr_date_exposure_dict, num_simulations_FX, directory)
    else:
        if directory != "combined":
            get_raw_prices(directory, output_type, False)
        else:
            raise Exception("Directory flag 'combined' can only be run with an output type of 3")


if __name__ == '__main__':
    parser = ArgumentParser(
        description="graph_simulations -d DATA_DIRECTORY -t TYPE_OUTPUT")
    parser.add_argument('-d', '--directory', help="Directory with all the data", required=True)
    parser.add_argument('-t', '--type', help="Type of output to produce", required=True, type=int)
    sysargs = parser.parse_args()
    main(sysargs.directory, sysargs.type)